import openpyxl

file_path = 'test runs_filtered_1_13_2026_4_41_12_PM.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print('Sheet name:', ws.title)
print('\nFirst 15 rows:')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 15:
        print(f'Row {i}: {row}')
    else:
        break

print('\nColumn headers:')
headers = [cell.value for cell in ws[1]]
for idx, header in enumerate(headers, 1):
    print(f'Column {idx}: {header}')
